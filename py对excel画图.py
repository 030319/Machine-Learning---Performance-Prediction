from openpyxl import load_workbook
from openpyxl.chart import (
    RadarChart,
    Reference,
)
import numpy as np

def read_all_data(worksheet):
    print("read_all_data...")
    data_list = []
    for row in worksheet.values:
        for value in row:
            data_list.append(value)
    print("data_list",data_list)
    dl = data_list
    return dl

def caculate_frequence(dl):
    print("caculate_frequence...")
    data_small_large = sorted(dl)
    #print("data",data)
    unique_data = np.unique(data_small_large)
    print("unique_data",unique_data)
    data_times = []
    for i in unique_data:
        data_times.append(data_small_large.count(i))
    print("resdata",resdata)
    times_total = sum(data_times)
    print("times_total",times_total)
    data_freq = []
    for i in data_times:
        freq_i = i / times_total
        data_freq.append(freq_i)
    df = data_freq
    print("df",df)
    return df

def draw_picture(df,dl):
    print("draw_picture...")
    for row in df:
        ws.append(row)
    chart = RadarChart()
    chart.type = "filled"
    #labels = Reference(ws, min_col=1, min_row=2, max_row=13)
    data = Reference(ws, min_col=1, max_col=1, min_row=1, max_row=len(dl))
    chart.add_data(data, titles_from_data=False)
    #chart.set_categories(labels)
    chart.style = 26
    chart.title = "wind direction frequence"
    chart.y_axis.delete = True
    ws.add_chart(chart, "B1")
    wb.save(r"路径")#放所需要保存的路径

def no_none(dl):
    print("no_none...")
    for item in dl[:]:
        if item == None:
            dl.remove(item)
    return dl

if __name__=="__main__":
    file_name = r'路径' + str(78) + '.xlsx'   #  读取数据的文件名
    wb = load_workbook(file_name)  #  加载工作本
    #print(wb.sheetnames)
    ws = wb.active  #  获取Sheet1
    data_list = read_all_data(ws)  #  读取所有的data
    data_list = no_none(data_list)  #  处理NaN(空值)
    data_freq = caculate_frequence(data_list)  #  计算频率
    draw_picture(data_freq,data_list)  #  画图

